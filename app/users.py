from datetime import datetime, timedelta
from io import BytesIO

from openpyxl.workbook import Workbook
from pymysql import IntegrityError

from app import app
import pymysql
from flask_wtf.csrf import CSRFProtect
import pandas as pd
from flask import request, flash, redirect, url_for, render_template, Response
import openpyxl
import calendar

csrf = CSRFProtect(app)


def db_connection():
    conn = pymysql.connect(host=app.config["DB_HOST"], user=app.config["DB_USERNAME"],
                           password=app.config["DB_PASSWORD"],
                           database=app.config["DB_NAME"])
    cursor = conn.cursor()
    return conn, cursor


def allowed_file(filename):
    return ('.' in filename and
            filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS'])


@app.route('/list_upload/<list_category>', methods=['GET', 'POST'])
def list_upload(list_category):
    if request.method == 'POST':
        conn, cursor = db_connection()
        file = request.files['list_file']
        if file.filename == '':
            flash('No selected file', 'warning')
            return redirect(url_for('list_upload', list_category=list_category))
        elif file and allowed_file(file.filename):
            try:
                file_data = pd.read_excel(file, sheet_name=None, engine='openpyxl', skiprows=5)
                if list_category == 'Attendance':
                    # Required columns
                    required_columns = {'Staff No.', 'Payment Rate'}
                    # Iterate over each sheet
                    for sheet_name, data in file_data.items():
                        # Check for missing columns
                        uploaded_columns = set(data.columns)
                        missing_columns = required_columns - uploaded_columns

                        if not missing_columns:
                            days_data_columns = [4, 5, 6, 7, 8, 9, 10]
                            # Iterate over each row in the sheet
                            for index, row in data.iloc[:-1].iterrows():
                                user_id = row['Staff No.']
                                payment_rate = row['Payment Rate']
                                # Now inner loop for extra columns
                                for col_index in days_data_columns:
                                    date = data.columns[col_index]  # header at that index
                                    cell_value = row.iloc[col_index]  # row value at that index
                                    status = 'X' if pd.isna(cell_value) else cell_value  # replace empty cells with 'X'

                                    # Insert into attendance table
                                    cursor.execute(
                                        "INSERT INTO attendance_records(user_id, date, payment_rate, status) "
                                        "VALUES(%s,%s,%s,%s)", (user_id, date, payment_rate, status)
                                    )
                            conn.commit()
                            flash("Attendance list uploaded successfully", 'success')
                        else:
                            missing_columns_str = ', '.join(missing_columns)
                            flash(f"The following columns are missing from {sheet_name} sheet: {missing_columns_str}. "
                                  f"Please confirm and re-upload.", 'danger')
                elif list_category == 'Payroll Summary':
                    date = request.form['date']
                    date_obj = datetime.strptime(date, "%Y-%m-%d")
                    # ISO calendar returns a tuple: (year, week_number, weekday)
                    iso_year, iso_week, iso_weekday = date_obj.isocalendar()

                    # Required columns
                    required_columns = {'Staff No.', 'Tips & Incentives', 'Gross', 'Shif', 'Nssf', 'Housing Levy',
                                        'Advances', 'Overpayment', 'Pending Bills', 'Total Dedution',
                                        'Housing Levy Refund', 'Paye',
                                        }
                    # Iterate over each sheet
                    for sheet_name, data in file_data.items():
                        # Check for missing columns
                        uploaded_columns = set(data.columns)
                        missing_columns = required_columns - uploaded_columns
                        if not missing_columns:
                            # Iterate over each row in the sheet
                            for index, row in data.iloc[:-1].iterrows():
                                # print(index)
                                user_id = row['Staff No.']
                                tips = row['Tips & Incentives']
                                gross = row['Gross']
                                paye = row['Paye']
                                shif = row['Shif']
                                nssf = row['Nssf']
                                housing_levy = row['Housing Levy']
                                advances = row['Advances']
                                overpayment = row['Overpayment']
                                pending_bills = row['Pending Bills']
                                total_deduction = row['Total Dedution']
                                housing_levy_refund = row['Housing Levy Refund']

                                # Insert into attendance table
                                cursor.execute("insert into payroll_summary (staff_no, tips, gross, paye, shif, nssf, "
                                               "housing_levy, advances, overpayment, pending_bills, total_deduction, "
                                               "housing_levy_refund, week, year)"
                                               "values(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                                               (user_id, tips, gross, paye, shif, nssf, housing_levy, advances,
                                                overpayment, pending_bills, total_deduction, housing_levy_refund,
                                                iso_week, iso_year))
                            conn.commit()
                            flash("list uploaded successfully", 'success')
                        else:
                            missing_columns_str = ', '.join(missing_columns)
                            flash(f"The following columns are missing from {sheet_name} sheet: {missing_columns_str}. "
                                  f"Please confirm and re-upload.", 'danger')
            except IntegrityError as e:
                error_code = e.args[0]
                conn.rollback()
                if error_code == 1062:  # MySQL error code for duplicate entry
                    flash(f"Duplicate entry error: ",
                          'danger')
                else:
                    flash(f"A DB error has occurred: {e}", 'danger')
            except Exception as e:
                flash(f"{e}", 'danger')
            finally:
                cursor.close()
                conn.close()
                return redirect(url_for('list_upload', list_category=list_category))
    return render_template('users.html', list_category=list_category)


def get_payroll_summary(year, week, unit):
    conn, cursor = db_connection()
    try:
        # ISO week: Sunday as first day
        first_day_of_the_week = datetime.fromisocalendar(year, week, 7)
        last_day_of_the_week = first_day_of_the_week + timedelta(days=6)

        # check whether the week crosses 2 different months
        crosses_month = first_day_of_the_week.month != last_day_of_the_week.month
        last_day_formated = last_day_of_the_week.strftime('%e %B %Y')
        # check if anything is passed in the units parameter
        if unit:
            # if it is passed, the data requested will be filtered to that specific unit
            operator = "="
            value = unit
        else:
            # if not, it will pull data of all the units
            operator = "like"
            value = "%"
        if crosses_month:
            # Get the month of that first day
            month = first_day_of_the_week.month

            # Get the last day of that month (date)
            last_day_of_month = calendar.monthrange(year, month)[1]  # returns (weekday_of_first_day, num_days_in_month)

            # Construct the date of the last day
            last_day_date = datetime(year, month, last_day_of_month)
            first_day_of_following_month = datetime(year, month + 1, 1)
            # compute the first range...
            cursor.execute("select '1;3;6', "
                           f"IF(month(%s) > 9 and month(%s) < 13, "
                           f"concat(year(%s) + 1, '/', lpad(month(%s) %% 9, 3, '0'))"
                           f", "
                           f"concat(year(%s), '/', lpad(month(%s) + 3, 3, '0'))) as period, "
                           "date_format(%s, '%%m/%%e/%%Y'), accounts.account_code, attendance_records.user_id, name, "
                           "sum(payment_rate), journal_marker, property_code, department, sub_department "
                           "from attendance_records "
                           "join casuals on attendance_records.user_id = casuals.user_id "
                           "join accounts_mapping on casuals.unit = accounts_mapping.unit_id "
                           "join accounts on accounts_mapping.account_id = accounts.account_id "
                           "join journals on journals.journal_id = accounts_mapping.journal_id "
                           "join `e-tamarind`.units on accounts_mapping.unit_id = units.unit_id "
                           "where status = %s and date between %s and %s and account_name = %s "
                           f"and units.unit_id {operator} %s "
                           "group by user_id",
                           (*[last_day_date] * 7, 'P', first_day_of_the_week, last_day_date, 'Basic Salary', value))
            first_range = cursor.fetchall()
            # required date format = 'date month_name 2025'
            sum_of_first_range = [last_day_date.strftime('%e %B %Y'), sum(row[6] for row in first_range)]
            # compute the second range...
            cursor.execute("select '1;3;6', "
                           f"IF(month(%s) > 9 and month(%s) < 13, "
                           f"concat(year(%s) + 1, '/', lpad(month(%s) %% 9, 3, '0'))"
                           f", "
                           f"concat(year(%s), '/', lpad(month(%s) + 3, 3, '0'))) as period, "
                           "date_format(%s, '%%m/%%e/%%Y'), accounts.account_code, attendance_records.user_id, name, "
                           "sum(payment_rate), journal_marker, property_code, department, sub_department "
                           "from attendance_records "
                           "join casuals on attendance_records.user_id = casuals.user_id "
                           "join casual_payments.accounts_mapping on casuals.unit = accounts_mapping.unit_id "
                           "join accounts on accounts_mapping.account_id = accounts.account_id "
                           "join journals on journals.journal_id = accounts_mapping.journal_id "
                           "join `e-tamarind`.units on accounts_mapping.unit_id = units.unit_id "
                           "where status = %s and date between %s and %s and account_name = %s "
                           f"and units.unit_id {operator} %s "
                           "group by user_id ",
                           (*[last_day_of_the_week] * 7, 'P', first_day_of_following_month, last_day_of_the_week,
                            'Basic Salary', value))
            second_range = cursor.fetchall()
            # required date format = 'date month_name 2025'
            sum_of_second_range = sum(row[6] for row in second_range)
            data = {
                'crosses_month': crosses_month,
                'first_range': first_range,
                'sum_of_first_range': sum_of_first_range,
                'second_range': second_range,
                'sum_of_second_range': sum_of_second_range,
                'last_day_formated': last_day_formated
            }
        else:
            cursor.execute("select '1;3;6', "
                           f"IF(month(%s) > 9 and month(%s) < 13, "
                           f"concat(year(%s) + 1, '/', lpad(month(%s) %% 9, 3, '0'))"
                           f", "
                           f"concat(year(%s), '/', lpad(month(%s) + 3, 3, '0'))) as period, "
                           "date_format(%s, '%%m/%%e/%%Y'), accounts.account_code, attendance_records.user_id, name, "
                           "sum(payment_rate), journal_marker, property_code, department, sub_department "
                           "from attendance_records "
                           "join casuals on attendance_records.user_id = casuals.user_id "
                           "join casual_payments.accounts_mapping on casuals.unit = accounts_mapping.unit_id "
                           "join accounts on accounts_mapping.account_id = accounts.account_id "
                           "join journals on journals.journal_id = accounts_mapping.journal_id "
                           "join `e-tamarind`.units on accounts_mapping.unit_id = units.unit_id "
                           "where status = %s and date between %s and %s and account_name = %s "
                           f"and units.unit_id {operator} %s "
                           "group by user_id ",
                           (*[last_day_of_the_week] * 7, 'P', first_day_of_the_week, last_day_of_the_week,
                            'Basic Salary', value))
            all_range = cursor.fetchall()
            sum_of_all_ranges = sum(row[6] for row in all_range)
            data = {
                'crosses_month': crosses_month,
                'all_range': all_range,
                'sum_of_all_ranges': sum_of_all_ranges,
                'last_day_formated': last_day_formated
            }
        # create empty dictionary to hold the respective data
        deductions_data = {}
        aggregated_data = {}
        # select available account names from db, except basic salary account because its calculation is done on
        # the queries above
        cursor.execute('select distinct account_name from accounts where account_name != %s', 'Basic Salary')
        accounts = cursor.fetchall()
        # hard code a list of the columns needed from the payroll summary table
        column_headers = ['tips', 'shif', 'nssf', 'housing_levy', 'advances', 'paye', 'pending_bills']
        # arrange the headers in alphabetical order in order to match the accounts order
        column_headers.sort()
        # using zip function, match each account to a column header in order to extract data
        for (account,), column_header in zip(accounts, column_headers):
            cursor.execute("select '1;3;6', "
                           f"IF(month(%s) > 9 and month(%s) < 13, "
                           f"concat(year(%s) + 1, '/', lpad(month(%s) %% 9, 3, '0'))"
                           f", "
                           f"concat(year(%s), '/', lpad(month(%s) + 3, 3, '0'))) as period, "
                           f" date_format(%s, '%%m/%%e/%%Y'), accounts.account_code, staff_no, name, {column_header}, "
                           f"journal_marker, property_code, department, sub_department "
                           "from payroll_summary "
                           "join casuals on staff_no = user_id "
                           "join casual_payments.accounts_mapping on casuals.unit = accounts_mapping.unit_id "
                           "join accounts on accounts_mapping.account_id = accounts.account_id "
                           "join journals on journals.journal_id = accounts_mapping.journal_id "
                           "join `e-tamarind`.units on accounts_mapping.unit_id = units.unit_id "
                           f"where week = %s and year = %s and account_name = %s and units.unit_id {operator} %s",
                           (*[last_day_of_the_week] * 7, week, year, account, value))
            extracted_data = cursor.fetchall()
            # create a dictionary to hold the deductions data with the account name as the key and the data
            # under that account as the value
            deductions = {
                f'{account}': extracted_data
            }
            # create another dictionary to hold the aggregated values of each account name
            deductions_sums = {
                f'{account}': [extracted_data[0][0], extracted_data[0][1], extracted_data[0][2],
                               sum(row[6] for row in extracted_data)]
            }
            # update the empty dictionaries from before with the data that we just created above
            deductions_data.update(deductions)
            aggregated_data.update(deductions_sums)
        return data, deductions_data, aggregated_data
    except Exception as e:
        raise e
    finally:
        cursor.close()
        conn.close()


@app.route('/payroll_summary', methods=['GET', 'POST'])
def payroll_summary():
    conn, cursor = db_connection()
    cursor.execute('select unit_id, unit_name from `e-tamarind`.units')
    unit_data = cursor.fetchall()
    cursor.close()
    conn.close()
    if request.method == 'POST':
        # Get year and week from the query parameters (sent from the form)
        year = int(request.form.get('year'))
        week = int(request.form.get('week'))
        unit = request.form.get('unit', "")

        # If no inputs provided, show page with form (GET request)
        if not year or not week:
            return render_template('payroll_summary.html')
        try:
            # get data from the function
            data, deductions_data, aggregated_data = get_payroll_summary(year, week, unit)
            # pass all the data to the payroll summary template for displaying
            return render_template('payroll_summary.html', data=data, deductions_data=deductions_data,
                                   aggregated_data=aggregated_data, unit_data=unit_data)
        except Exception as e:
            flash(f"An error occurred {e}", 'danger')
            return redirect(url_for('payroll_summary'))
    return render_template('payroll_summary.html', unit_data=unit_data)


@app.route('/generate_excel')
def generate_excel():
    #  connect to database
    conn, cursor = db_connection()
    # Get year and week from the arguments
    year = int(request.args.get('year'))
    week = int(request.args.get('week'))
    # get data from the function
    data, deductions_data, aggregated_data = get_payroll_summary(year, week)
    # Create a new Excel workbook
    wb = Workbook()
    # Remove the default sheet
    if "Sheet" in wb.sheetnames:
        default_sheet = wb["Sheet"]
        wb.remove(default_sheet)

    cursor.execute("select property_code, unit_name from `e-tamarind`.units where unit_name != %s ",
                   'Tamarind Village')
    units = cursor.fetchall()
    cursor.close()
    conn.close()
    for unit in units:
        property_code = unit[0]
        unit_name = unit[1]

        ws = wb.create_sheet(title=unit_name)

        # Write header to Excel
        headers = ['1', 'Period', 'Date', 'Account Code', 'Reference', 'Description', 'Base Amount', ' Dr/Cr',
                   'Property', 'Outlet', 'Sub Outlet', 'Journal Type']
        for col_index, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_index, value=header)
            # Fetch and write data to Excel
            # Write data for this property
            row_index = 2
            # Calculation of the sum of basic salary/gross pay
            # if data is in 2 different months
            if data['crosses_month']:
                # Filter rows that match this property_code
                property_rows_first = [row for row in data['first_range'] if row[8] == property_code]
                property_rows_second = [row for row in data['second_range'] if row[8] == property_code]
                # display sum of the first range
                if property_rows_first:
                    first_range_total = sum(row[6] for row in property_rows_first)
                    ws.cell(row=row_index, column=1, value=data['first_range'][0][0])
                    ws.cell(row=row_index, column=2, value=data['first_range'][0][1])
                    ws.cell(row=row_index, column=3, value=data['first_range'][0][2])
                    ws.cell(row=row_index, column=4, value=110383)
                    ws.cell(row=row_index, column=5, value=data['sum_of_first_range'][0])
                    ws.cell(row=row_index, column=6, value='M-pesa Payments')
                    ws.cell(row=row_index, column=7, value=first_range_total)
                    ws.cell(row=row_index, column=8, value='C')
                    ws.cell(row=row_index, column=9, value=property_code)
                    ws.cell(row=row_index, column=10, value='')
                    ws.cell(row=row_index, column=11, value='')
                    ws.cell(row=row_index, column=12, value='MPESA')
                    row_index += 1
                # display sum of the second range
                if property_rows_second:
                    second_range_total = sum(row[6] for row in property_rows_second)
                    ws.cell(row=row_index, column=1, value=data['second_range'][0][0])
                    ws.cell(row=row_index, column=2, value=data['second_range'][0][1])
                    ws.cell(row=row_index, column=3, value=data['second_range'][0][2])
                    ws.cell(row=row_index, column=4, value=110383)
                    ws.cell(row=row_index, column=5, value=data['last_day_formated'])
                    ws.cell(row=row_index, column=6, value='M-pesa Payments')
                    ws.cell(row=row_index, column=7, value=second_range_total)
                    ws.cell(row=row_index, column=8, value='C')
                    ws.cell(row=row_index, column=9, value=property_code)
                    ws.cell(row=row_index, column=10, value='')
                    ws.cell(row=row_index, column=11, value='')
                    ws.cell(row=row_index, column=12, value='MPESA')
                    row_index += 1
            # if data is in the same month
            else:
                # Filter rows that match this property_code
                property_rows_all = [row for row in data['all_range'] if row[8] == property_code]
                # display the sum
                if property_rows_all:
                    all_range_total = sum(row[6] for row in property_rows_all)
                    ws.cell(row=row_index, column=1, value=data['all_range'][0][0])
                    ws.cell(row=row_index, column=2, value=data['all_range'][0][1])
                    ws.cell(row=row_index, column=3, value=data['all_range'][0][2])
                    ws.cell(row=row_index, column=4, value=110383)
                    ws.cell(row=row_index, column=5, value=data['last_day_formated'])
                    ws.cell(row=row_index, column=6, value='M-pesa Payments')
                    ws.cell(row=row_index, column=7, value=all_range_total)
                    ws.cell(row=row_index, column=8, value='C')
                    ws.cell(row=row_index, column=9, value=property_code)
                    ws.cell(row=row_index, column=10, value='')
                    ws.cell(row=row_index, column=11, value='')
                    ws.cell(row=row_index, column=12, value='MPESA')
                    row_index += 1
            # loop through the deductions data to create sum of each deduction
            for key, values in deductions_data.items():
                # Filter rows belonging to this property
                deductions_data_rows = [row for row in values if row[8] == property_code]
                if deductions_data_rows:
                    total = sum(row[6] for row in deductions_data_rows)
                    ws.cell(row=row_index, column=1, value=values[0][0])
                    ws.cell(row=row_index, column=2, value=values[0][1])
                    ws.cell(row=row_index, column=3, value=values[0][2])
                    ws.cell(row=row_index, column=4, value=110383)
                    ws.cell(row=row_index, column=5, value=data['last_day_formated'])
                    ws.cell(row=row_index, column=6, value=key)
                    ws.cell(row=row_index, column=7, value=total)
                    ws.cell(row=row_index, column=8, value='D')
                    ws.cell(row=row_index, column=9, value=property_code)
                    ws.cell(row=row_index, column=10, value='')
                    ws.cell(row=row_index, column=11, value='')
                    ws.cell(row=row_index, column=12, value='MPESA')
                    row_index += 1
            # if data is between 2 months
            if data['crosses_month']:
                # display the detailed section of the basic salary first range
                for item in data['first_range']:
                    if item[8] == property_code:  # property_code column
                        for col_index, value in enumerate(item[:11], start=1):
                            ws.cell(row=row_index, column=col_index, value=value)
                        ws.cell(row=row_index, column=12, value="MPESA")
                        row_index += 1
                # display the detailed section of the basic salary second range
                for item in data['second_range']:
                    if item[8] == property_code:  # property_code column
                        for col_index, value in enumerate(item[:11], start=1):
                            ws.cell(row=row_index, column=col_index, value=value)
                        ws.cell(row=row_index, column=12, value="MPESA")
                        row_index += 1
            # if it is all in one month
            else:
                # display the detailed section of the basic salary all range
                for item in data['all_range']:
                    if item[8] == property_code:  # property_code column
                        for col_index, value in enumerate(item[:11], start=1):
                            ws.cell(row=row_index, column=col_index, value=value)
                        ws.cell(row=row_index, column=12, value="MPESA")
                        row_index += 1
            # for loop to generate the detailed data of the deductions section (advances, housing levy, NSSF,
            #                 PAYE, Pending bills, SHIF, tips)
            for key, values in deductions_data.items():
                for row in values:
                    if row[8] == property_code:
                        for col_index, value in enumerate(row[:11], start=1):
                            ws.cell(row=row_index, column=col_index, value=value)
                        ws.cell(row=row_index, column=12, value="MPESA")
                        row_index += 1

    # Save the workbook
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    # Create a downloadable response
    return Response(
        output.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":
                     f"attachment; filename=Payroll Summary.xlsx"}
    )


@app.route('/')
def home():
    return render_template('home.html')
